from django.shortcuts import render, HttpResponse, redirect
from openpyxl.reader.excel import load_workbook
from user import models
from user.utils.pagination import Pagination
from user.utils.bootstrap import BootStrapModelForm


class UpModelForm(BootStrapModelForm):
    class Meta:
        model = models.ExcelInfo
        fields = ['file_type']
        # labels = {
        #     'file_type': '选择项目：'
        # }
        # widgets = {
        #     'file_type': forms.Select(attrs={"data-placeholder": "选择项目", }),
        # }

    # def __init__(self, *args, **kwargs):
    #     super(UpModelForm, self).__init__(*args, **kwargs)
    #     self.fields['file_type'].initial = '请选择项目...'

    # 根据excel中第一行的列名，将该行数据对应到ExcelInfo模型中相应的字段
    @staticmethod
    def parse_excel_row(header_row, data_row, file_type, file_title):
        headers = [c.value for c in header_row]
        data = {headers[i]: data_row[i].value for i in range(len(headers))}
        obj = models.ExcelInfo(file_type=file_type, file_title=file_title)
        for attr, verbose_name in [(f.name, f.verbose_name) for f in obj._meta.fields]:
            if verbose_name in headers:
                setattr(obj, attr, data[verbose_name])
        return obj


def multi_upload(request):
    form = UpModelForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            file_object = request.FILES.get('excel_file')
            if not file_object:
                return HttpResponse('请选择上传的文件！')  # excel_file为空，返回错误提示
            file_type = form.cleaned_data.get('file_type')  # 获取file_type的值
            if not file_type:
                return HttpResponse('请选择项目！')  # file_type为空，返回错误提示
            # 使用 openpyxl 进行读取文件处理
            wb = load_workbook(file_object)
            sheet = wb.worksheets[0]

            # 判断上传文件名不允许重复
            file_title = request.FILES['excel_file'].name
            if models.ExcelInfo.objects.filter(file_title=file_title).exists():
                return HttpResponse('该文件已经上传，请勿重复上传！')
            form.instance.file_title = file_title

            # 获取表头行和数据起始行
            header_row = sheet[1]
            data_start_row = 2

            # 逐行读取数据，并解析到ExcelInfo模型中
            for row in sheet.iter_rows(min_row=data_start_row):
                obj = UpModelForm.parse_excel_row(header_row, row, file_type, file_title)
                obj.file_title = file_title  # 将文件名赋值给file_title字段
                obj.save()

            return redirect("/document/list/")
    return redirect("/document/list/")

    # context = {
    #     'form': form
    # }
    # return render(request, 'multi_list.html', context)


def multi_list(request):
    if request.GET.get('query'):
        queryset = models.ExcelInfo.objects.filter(file_title__contains=request.GET.get('query'))
    else:
        queryset = models.ExcelInfo.objects.all()

    # 分页
    page_object = Pagination(request, queryset)

    context = {
        'queryset': page_object.page_queryset,
        'page_string': page_object.html(),
        'form': UpModelForm(),
        'search_data': request.GET.get('query', ''),
    }
    return render(request, 'multi_list.html', context)


def multi_add(request):
    pass
